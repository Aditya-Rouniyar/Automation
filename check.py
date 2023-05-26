from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

cloudengineer = './sheets/cloudengineer.xlsx'
def test():
    wb = load_workbook(cloudengineer)
    ws = wb.active
    print(list(ws.values))
    #ws.append(('zoom',0,0,0,0,0,0,0))
    #wb.save('./sheets/devop.xlsx')
test()